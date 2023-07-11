import csv
import os
import random
import openpyxl

def read_xlsx_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def read_csv_file(file_path):
    data = []
    with open(file_path, 'r') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            quantity = row['Quantity']
            manf = row['manf#']
            if manf:  # Skip rows where 'manf' value is empty
                data.append([manf, quantity])
    return data

def write_csv_file(file_path, BOM, Number_of_boards):

    directory = os.path.dirname(file_path)
    output_file = os.path.join(directory, 'BOM_FOUND.csv')

    with open(output_file, 'w', newline='') as file:
        writer = csv.writer(file)
        header = ['manf#', 'Quantity', 'Found']
        writer.writerow(list(header))

        for row in BOM:
            manf = row[0]
            quantity = int(row[1]) * int(Number_of_boards)
            found = row[2]
            writer.writerow([quantity, manf, found])  # Write the updated row

    print("The 'BOM_FOUND.csv' file has been created.")

def write_xlsx_file(file_path, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in data:
        sheet.append(row)
    workbook.save(file_path)
    print(f"The '{file_path}' file has been created.")


def main():
    # Prompt CSV BOM file path
    BOM_CSV_Path = input("Enter the local path of the BOM CSV file: ")
    BOM_CSV_Path = BOM_CSV_Path.strip("& '")

    # Prompt XLSX stock file path
    STOCK_XLSX_Path = input("Enter the local path of the STOCK XLSX file: ")
    STOCK_XLSX_Path = STOCK_XLSX_Path.strip("& '")

    # Prompt the number of fabricated boards
    # Number_of_boards = input("Enter the number of fabricated boards: ")

    # Read the BOM file and store the content of 'quantity' and 'manf' columns in a 2D array
    BOM_DATA = read_csv_file(BOM_CSV_Path)



    # Read the XLSX file and store its content in a 2D array
    STOCK_DATA = read_xlsx_file(STOCK_XLSX_Path)

    for row in STOCK_DATA:
        if isinstance(row[0], int):
            row[0] -= 1

    directory_STOCK = os.path.dirname(STOCK_XLSX_Path)
    output_file_STOCK = os.path.join(directory_STOCK, 'STOCK_UPDATED.xlsx')
    write_xlsx_file(output_file_STOCK, STOCK_DATA)

    # Print the content of the XLSX file
    for row in STOCK_DATA:
        print(row)

    # # Write the 2D array into a CSV file named 'BOM_FOUND' at the same path as the opened CSV file

    # write_csv_file(BOM_CSV_Path, BOM_DATA, Number_of_boards)

#**********************************************#
#                    main                      #
#**********************************************#
if __name__ == "__main__":
    main()
